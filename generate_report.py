#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generador de Informe de Soporte Técnico - AFJ Global
Basado en datos de Freshdesk
"""

import openpyxl
from datetime import datetime
from collections import Counter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
import io

# Configuración
EXCEL_FILE = 'reporte_freshdesk_AFJ_Global.xlsx'
OUTPUT_PDF = 'Informe_Soporte_Tecnico_AFJ_Global.pdf'
COMPANY_NAME = 'AFJ Global'
REPORT_DATE = datetime.now().strftime('%d de %B de %Y')
REPORT_VERSION = '1.0'

# Colores corporativos (similar a Consultame)
COLOR_PRIMARY = colors.HexColor('#667eea')
COLOR_SECONDARY = colors.HexColor('#764ba2')
COLOR_ALTA = colors.HexColor('#ff6b6b')
COLOR_MEDIA = colors.HexColor('#ffd93d')
COLOR_BAJA = colors.HexColor('#6bcf7f')

def load_excel_data():
    """Carga y analiza datos del Excel"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['Todos los Tickets']

    tickets = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticket = dict(zip(headers, row))
        tickets.append(ticket)

    return tickets

def analyze_data(tickets):
    """Analiza los datos de tickets"""
    total_tickets = len(tickets)

    # Por prioridad
    priority_count = Counter([t.get('priority', 'N/A') for t in tickets])

    # Por estado
    status_count = Counter()
    status_map = {
        '2': 'Abierto',
        '3': 'Pendiente',
        '4': 'Resuelto',
        '5': 'Cerrado'
    }
    for t in tickets:
        status = str(t.get('status', 'N/A'))
        status_name = status_map.get(status, f'Estado {status}')
        status_count[status_name] += 1

    # Por año
    year_count = Counter()
    for t in tickets:
        if t.get('created_at'):
            try:
                year = datetime.fromisoformat(str(t['created_at']).replace('Z', '+00:00')).year
                year_count[year] += 1
            except:
                pass

    # Top 10 asuntos
    subject_count = Counter()
    for t in tickets:
        subject = t.get('subject') or ''
        subject = subject.strip() if isinstance(subject, str) else ''
        if subject:
            subject_count[subject] += 1
    top_10_subjects = subject_count.most_common(10)

    # Clasificación por contenido (similar a la lógica del servidor)
    clasificacion = {'ALTA': 0, 'MEDIA': 0, 'BAJA': 0}

    keywords_alta = ['aws', 'alarm', 'no enciende', 'no prende', 'escritorio remoto',
                     'virus', 'malware', 'error servidor', 'afjlearning',
                     'sharepoint lentitud', 'moodle']
    keywords_media = ['outlook', 'correo', 'fundae', 'limpieza', 'buzón',
                      'antivirus', 'pst', 'revisar pc', 'sharepoint', 'spam',
                      'cambio licencia', 'acceso carpeta']
    keywords_baja = ['revisión windows', 'mensaje ausencia', 'alta', 'baja',
                     'redirección', 'instalación']

    for ticket in tickets:
        subject_val = ticket.get('subject') or ''
        desc_val = ticket.get('description') or ''
        subject = (str(subject_val) + ' ' + str(desc_val)).lower()

        if any(kw in subject for kw in keywords_alta):
            clasificacion['ALTA'] += 1
        elif any(kw in subject for kw in keywords_media):
            clasificacion['MEDIA'] += 1
        elif any(kw in subject for kw in keywords_baja):
            clasificacion['BAJA'] += 1
        else:
            clasificacion['BAJA'] += 1

    return {
        'total_tickets': total_tickets,
        'priority_count': priority_count,
        'status_count': status_count,
        'year_count': year_count,
        'top_10_subjects': top_10_subjects,
        'clasificacion': clasificacion
    }

def create_pie_chart(data_dict, title, width=300, height=200):
    """Crea un gráfico de dona/pie"""
    drawing = Drawing(width, height)

    pie = Pie()
    pie.x = 50
    pie.y = 50
    pie.width = 150
    pie.height = 150

    labels = list(data_dict.keys())
    values = list(data_dict.values())

    pie.data = values
    pie.labels = [f'{l}: {v}' for l, v in zip(labels, values)]

    # Colores
    colors_list = [COLOR_ALTA, COLOR_MEDIA, COLOR_BAJA, COLOR_PRIMARY,
                   COLOR_SECONDARY, colors.green, colors.orange, colors.red]
    pie.slices.strokeWidth = 0.5
    for i, color in enumerate(colors_list[:len(values)]):
        pie.slices[i].fillColor = color

    drawing.add(pie)
    return drawing

def create_bar_chart(data_dict, title, width=400, height=200):
    """Crea un gráfico de barras"""
    drawing = Drawing(width, height)

    bar = VerticalBarChart()
    bar.x = 50
    bar.y = 50
    bar.height = 125
    bar.width = 300

    labels = list(data_dict.keys())
    values = list(data_dict.values())

    bar.data = [values]
    bar.categoryAxis.categoryNames = labels

    bar.bars[0].fillColor = COLOR_PRIMARY
    bar.valueAxis.valueMin = 0
    bar.valueAxis.valueMax = max(values) * 1.2 if values else 100

    drawing.add(bar)
    return drawing

class ReportTemplate:
    """Clase para generar el PDF con estilos consistentes"""

    def __init__(self, filename):
        self.filename = filename
        self.doc = SimpleDocTemplate(filename, pagesize=A4,
                                      rightMargin=72, leftMargin=72,
                                      topMargin=72, bottomMargin=72)
        self.styles = getSampleStyleSheet()
        self.story = []

        # Estilos personalizados
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=COLOR_PRIMARY,
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='CustomHeading2',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=COLOR_SECONDARY,
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='CustomBody',
            parent=self.styles['BodyText'],
            fontSize=11,
            alignment=TA_JUSTIFY,
            spaceAfter=12
        ))

    def add_cover_page(self):
        """Crea la página de portada"""
        # Título principal
        self.story.append(Spacer(1, 2*inch))

        title = Paragraph(
            f"<b>INFORME DE SOPORTE TÉCNICO</b>",
            self.styles['CustomTitle']
        )
        self.story.append(title)
        self.story.append(Spacer(1, 0.3*inch))

        subtitle = Paragraph(
            f"<b>Análisis de Tickets de Mesa de Ayuda</b>",
            ParagraphStyle(
                name='Subtitle',
                fontSize=16,
                textColor=COLOR_SECONDARY,
                alignment=TA_CENTER,
                spaceAfter=50
            )
        )
        self.story.append(subtitle)
        self.story.append(Spacer(1, 1*inch))

        # Información del cliente
        info_data = [
            ['Cliente:', COMPANY_NAME],
            ['Fecha:', REPORT_DATE],
            ['Versión:', REPORT_VERSION],
            ['Elaborado por:', 'Mesa de Ayuda TI']
        ]

        info_table = Table(info_data, colWidths=[2*inch, 3.5*inch])
        info_table.setStyle(TableStyle([
            ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 12),
            ('FONT', (1, 0), (1, -1), 'Helvetica', 12),
            ('TEXTCOLOR', (0, 0), (0, -1), COLOR_PRIMARY),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))

        self.story.append(info_table)
        self.story.append(PageBreak())

    def add_executive_summary(self, data):
        """Agrega el resumen ejecutivo"""
        self.story.append(Paragraph("<b>RESUMEN EJECUTIVO</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        summary_text = f"""
        Durante el período analizado, la Mesa de Ayuda de {COMPANY_NAME} ha gestionado un total de
        <b>{data['total_tickets']} tickets de soporte técnico</b>. Este informe presenta un análisis
        detallado del rendimiento del servicio, la distribución de incidencias por criticidad,
        y las principales áreas de atención requeridas por los usuarios.
        """

        self.story.append(Paragraph(summary_text, self.styles['CustomBody']))
        self.story.append(Spacer(1, 0.3*inch))

        # Tabla de resumen
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total de Tickets', str(data['total_tickets'])],
            ['Tickets Cerrados', str(data['status_count'].get('Cerrado', 0))],
            ['Tickets Abiertos', str(data['status_count'].get('Abierto', 0))],
            ['Criticidad Alta', str(data['clasificacion']['ALTA'])],
            ['Criticidad Media', str(data['clasificacion']['MEDIA'])],
            ['Criticidad Baja', str(data['clasificacion']['BAJA'])],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        self.story.append(summary_table)
        self.story.append(Spacer(1, 0.3*inch))

    def add_classification_analysis(self, data):
        """Agrega análisis de clasificación por criticidad"""
        self.story.append(PageBreak())
        self.story.append(Paragraph("<b>ANÁLISIS POR CRITICIDAD</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        text = """
        Los tickets han sido clasificados según su criticidad en tres niveles: ALTA (incidentes críticos
        que afectan servicios esenciales), MEDIA (problemas importantes que requieren atención),
        y BAJA (solicitudes de mantenimiento rutinario).
        """

        self.story.append(Paragraph(text, self.styles['CustomBody']))
        self.story.append(Spacer(1, 0.2*inch))

        # Gráfico de clasificación
        clasificacion_chart = create_pie_chart(data['clasificacion'], 'Distribución por Criticidad')
        self.story.append(clasificacion_chart)
        self.story.append(Spacer(1, 0.3*inch))

        # Detalles por criticidad
        criticidad_data = [
            ['Criticidad', 'Cantidad', 'Porcentaje'],
            ['🔴 ALTA (Crítico)',
             str(data['clasificacion']['ALTA']),
             f"{data['clasificacion']['ALTA']/data['total_tickets']*100:.1f}%"],
            ['🟡 MEDIA (Importante)',
             str(data['clasificacion']['MEDIA']),
             f"{data['clasificacion']['MEDIA']/data['total_tickets']*100:.1f}%"],
            ['🟢 BAJA (Normal)',
             str(data['clasificacion']['BAJA']),
             f"{data['clasificacion']['BAJA']/data['total_tickets']*100:.1f}%"],
        ]

        criticidad_table = Table(criticidad_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        criticidad_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_SECONDARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        self.story.append(criticidad_table)
        self.story.append(Spacer(1, 0.3*inch))

        # Descripción de categorías
        categories_text = f"""
        <b>Categorías de Criticidad:</b><br/>
        <br/>
        <b>🔴 ALTA (Crítico):</b> Incluye incidentes relacionados con AWS, alarmas de sistema,
        equipos que no encienden, problemas de escritorio remoto, virus/malware, errores de servidor,
        problemas con AFJLearning, lentitud en SharePoint y problemas con Moodle.
        Total: {data['clasificacion']['ALTA']} tickets.<br/>
        <br/>
        <b>🟡 MEDIA (Importante):</b> Abarca problemas con Outlook, correos electrónicos,
        acceso a FUNDAE, limpieza de buzones, alertas de antivirus, archivos PST, revisiones de PC,
        problemas generales de SharePoint, correos spam, cambios de licencia y accesos a carpetas.
        Total: {data['clasificacion']['MEDIA']} tickets.<br/>
        <br/>
        <b>🟢 BAJA (Normal):</b> Comprende revisiones rutinarias de Windows, mensajes de ausencia,
        altas y bajas de usuarios, redirección de correos e instalaciones de software estándar.
        Total: {data['clasificacion']['BAJA']} tickets.
        """

        self.story.append(Paragraph(categories_text, self.styles['CustomBody']))

    def add_status_analysis(self, data):
        """Agrega análisis por estado"""
        self.story.append(PageBreak())
        self.story.append(Paragraph("<b>ANÁLISIS POR ESTADO</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        # Gráfico de estados
        status_chart = create_pie_chart(data['status_count'], 'Distribución por Estado')
        self.story.append(status_chart)
        self.story.append(Spacer(1, 0.3*inch))

        # Tabla de estados
        status_data = [['Estado', 'Cantidad', 'Porcentaje']]
        for status, count in data['status_count'].items():
            percentage = f"{count/data['total_tickets']*100:.1f}%"
            status_data.append([status, str(count), percentage])

        status_table = Table(status_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        self.story.append(status_table)

    def add_top_tickets(self, data):
        """Agrega análisis de tickets más recurrentes"""
        self.story.append(PageBreak())
        self.story.append(Paragraph("<b>TOP 10 TICKETS MÁS RECURRENTES</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        text = """
        Los siguientes tickets representan las incidencias y solicitudes más frecuentes
        registradas en el período analizado. Identificar estos patrones permite optimizar
        recursos y prevenir incidencias recurrentes.
        """

        self.story.append(Paragraph(text, self.styles['CustomBody']))
        self.story.append(Spacer(1, 0.2*inch))

        # Tabla de top 10
        top_data = [['#', 'Asunto', 'Cantidad']]
        for i, (subject, count) in enumerate(data['top_10_subjects'], 1):
            # Truncar asunto si es muy largo
            subject_short = subject[:60] + '...' if len(subject) > 60 else subject
            top_data.append([str(i), subject_short, str(count)])

        top_table = Table(top_data, colWidths=[0.5*inch, 4*inch, 1*inch])
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_SECONDARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        self.story.append(top_table)

    def add_recommendations(self, data):
        """Agrega recomendaciones"""
        self.story.append(PageBreak())
        self.story.append(Paragraph("<b>RECOMENDACIONES</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        # Alta prioridad
        if data['clasificacion']['ALTA'] > 0:
            self.story.append(Paragraph("<b>🔴 PRIORIDAD ALTA:</b>", self.styles['Heading3']))
            rec_alta = f"""
            Se han identificado {data['clasificacion']['ALTA']} tickets de criticidad alta.
            Se recomienda:<br/>
            • Establecer un protocolo de respuesta inmediata para incidentes críticos<br/>
            • Implementar monitoreo proactivo de servicios esenciales (AWS, SharePoint, Moodle)<br/>
            • Revisar y actualizar la documentación de procedimientos de emergencia
            """
            self.story.append(Paragraph(rec_alta, self.styles['CustomBody']))
            self.story.append(Spacer(1, 0.2*inch))

        # Media prioridad
        self.story.append(Paragraph("<b>🟡 PRIORIDAD MEDIA:</b>", self.styles['Heading3']))
        rec_media = f"""
        Con {data['clasificacion']['MEDIA']} tickets de prioridad media, se sugiere:<br/>
        • Crear guías de autoayuda para problemas comunes de Outlook y correo<br/>
        • Implementar limpieza automática de buzones para prevenir problemas de espacio<br/>
        • Capacitar a usuarios en el uso correcto del antivirus y gestión de archivos PST
        """
        self.story.append(Paragraph(rec_media, self.styles['CustomBody']))
        self.story.append(Spacer(1, 0.2*inch))

        # Mantenimiento
        self.story.append(Paragraph("<b>🟢 MANTENIMIENTO:</b>", self.styles['Heading3']))
        rec_baja = f"""
        Para los {data['clasificacion']['BAJA']} tickets de mantenimiento rutinario:<br/>
        • Automatizar la programación de revisiones de Windows<br/>
        • Implementar plantillas para mensajes de ausencia<br/>
        • Agilizar el proceso de altas/bajas mediante formularios estandarizados
        """
        self.story.append(Paragraph(rec_baja, self.styles['CustomBody']))

    def add_conclusions(self, data):
        """Agrega conclusiones"""
        self.story.append(PageBreak())
        self.story.append(Paragraph("<b>CONCLUSIONES</b>", self.styles['CustomHeading2']))
        self.story.append(Spacer(1, 0.2*inch))

        conclusions_text = f"""
        El análisis de los {data['total_tickets']} tickets procesados revela un servicio
        de Mesa de Ayuda activo y funcional, con la mayoría de los tickets
        ({data['status_count'].get('Cerrado', 0)}) resueltos satisfactoriamente.
        <br/><br/>
        La distribución de tickets por criticidad muestra que el {data['clasificacion']['BAJA']/data['total_tickets']*100:.1f}%
        corresponde a mantenimiento rutinario, mientras que solo el {data['clasificacion']['ALTA']/data['total_tickets']*100:.1f}%
        requiere atención crítica inmediata. Esto indica una gestión preventiva efectiva.
        <br/><br/>
        Los tickets más recurrentes están relacionados principalmente con revisiones de Windows y
        gestión de buzones de correo, sugiriendo oportunidades claras para automatización y mejora de procesos.
        <br/><br/>
        Se recomienda mantener el nivel de servicio actual, implementar las mejoras sugeridas en
        la sección de recomendaciones, y continuar con el monitoreo periódico del rendimiento
        del servicio de soporte técnico.
        """

        self.story.append(Paragraph(conclusions_text, self.styles['CustomBody']))
        self.story.append(Spacer(1, 0.5*inch))

        # Firma
        signature = Paragraph(
            f"<b>Mesa de Ayuda TI<br/>{COMPANY_NAME}<br/>{REPORT_DATE}</b>",
            ParagraphStyle(
                name='Signature',
                fontSize=11,
                alignment=TA_RIGHT,
                textColor=COLOR_PRIMARY
            )
        )
        self.story.append(signature)

    def build(self):
        """Construye el PDF"""
        self.doc.build(self.story)

def main():
    """Función principal"""
    print("📊 Generando Informe de Soporte Técnico - AFJ Global")
    print("=" * 60)

    # Cargar datos
    print("📁 Cargando datos del Excel...")
    tickets = load_excel_data()

    # Analizar datos
    print("🔍 Analizando tickets...")
    data = analyze_data(tickets)

    print(f"✅ Total de tickets: {data['total_tickets']}")
    print(f"✅ Clasificación: ALTA={data['clasificacion']['ALTA']}, " +
          f"MEDIA={data['clasificacion']['MEDIA']}, BAJA={data['clasificacion']['BAJA']}")

    # Generar PDF
    print("\n📄 Generando PDF...")
    report = ReportTemplate(OUTPUT_PDF)

    report.add_cover_page()
    report.add_executive_summary(data)
    report.add_classification_analysis(data)
    report.add_status_analysis(data)
    report.add_top_tickets(data)
    report.add_recommendations(data)
    report.add_conclusions(data)

    report.build()

    print(f"\n✅ ¡Informe generado exitosamente!")
    print(f"📍 Ubicación: {OUTPUT_PDF}")
    print("=" * 60)

if __name__ == '__main__':
    main()
